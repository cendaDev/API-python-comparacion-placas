from fastapi import FastAPI, UploadFile, File
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi import HTTPException
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import csv
#python -m uvicorn main:app --reload
#pip install fastapi uvicorn pandas openpyxl
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def detectar_separador_robusto(content: bytes) -> str:
    """Detecta el separador y valida que las columnas se lean correctamente."""
    sample = content.decode('utf-8', errors='ignore')
    posibles = [',', ';', '\t', '|']

    try:
        dialect = csv.Sniffer().sniff(sample[:2048])
        return dialect.delimiter
    except Exception:
        pass  # continúa con detección manual

    for sep in posibles:
        try:
            df_test = pd.read_csv(BytesIO(content), sep=sep, nrows=5)
            if len(df_test.columns) > 1:
                return sep
        except Exception:
            continue
    return ','  # fallback

@app.post("/comparar-excel")
async def comparar_archivos(file_actual: UploadFile = File(...), file_pasado: UploadFile = File(...)):
    content1 = await file_actual.read()
    content2 = await file_pasado.read()

    # Detectar extensión
    ext1 = file_actual.filename.split(".")[-1].lower()
    ext2 = file_pasado.filename.split(".")[-1].lower()

    # Cargar con pandas según el tipo
    # Cargar con pandas según el tipo
    if ext1 == "csv":
        sep1 = detectar_separador_robusto(content1)
        df_actual = pd.read_csv(BytesIO(content1), sep=sep1, encoding='utf-8-sig', on_bad_lines='skip')
        if len(df_actual.columns) == 1:  # si sigue detectando solo 1 columna
            df_actual = pd.read_csv(BytesIO(content1), sep=';', encoding='utf-8-sig', on_bad_lines='skip')
    elif ext1 in ["xls", "xlsx"]:
        df_actual = pd.read_excel(BytesIO(content1))
    else:
        raise HTTPException(status_code=400, detail=f"Formato no soportado: {file_actual.filename}")

    if ext2 == "csv":
        sep2 = detectar_separador_robusto(content2)
        df_pasado = pd.read_csv(BytesIO(content2), sep=sep2, encoding='utf-8-sig', on_bad_lines='skip')
        if len(df_pasado.columns) == 1:
            df_pasado = pd.read_csv(BytesIO(content2), sep=';', encoding='utf-8-sig', on_bad_lines='skip')
    elif ext2 in ["xls", "xlsx"]:
        df_pasado = pd.read_excel(BytesIO(content2))
    else:
        raise HTTPException(status_code=400, detail=f"Formato no soportado: {file_pasado.filename}")

    # Normalizar nombres de columnas
    df_actual.columns = df_actual.columns.str.lower().str.strip()
    df_pasado.columns = df_pasado.columns.str.lower().str.strip()

    print(df_actual.columns)

    # Asegurar que existen 'placa' y 'tipo revision'
    if "placa" not in df_actual.columns or "tipo revision" not in df_actual.columns:
        return {"error": "El archivo actual no tiene columnas 'placa' y 'tipo revision' 1"}
    if "placa" not in df_pasado.columns or "tipo revision" not in df_pasado.columns:
        return {"error": "El archivo pasado no tiene columnas 'placa' y 'tipo revision' 2"}

    # Comparar
    placas_actual = set(df_actual["placa"].astype(str).str.strip())
    placas_pasado = set(df_pasado["placa"].astype(str).str.strip())

    # 1. Ya vino o Primera vez
    df_actual["estado"] = df_actual["placa"].apply(
        lambda p: "Ya vino" if p in placas_pasado else "Primera vez"
    )

    # 2. No vino → de las del pasado que no están en el actual
    df_no_vino = df_pasado[~df_pasado["placa"].isin(placas_actual)].copy()
    df_no_vino["estado"] = "No vino"

    # Unir todo
    df_final = pd.concat([df_actual, df_no_vino], ignore_index=True)

    # Guardar a Excel en memoria
    output = BytesIO()
    df_final.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)

    # Reabrir con openpyxl para aplicar colores
    wb = load_workbook(output)
    ws = wb.active

    # Buscar índice de la columna "estado"
    estado_col = None
    for i, col in enumerate(ws[1], 1):
        if col.value == "estado":
            estado_col = i
            break

    # Colores
    fill_verde = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde
    fill_rojo = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")  # Rojo
    fill_azul = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # Azul

    # Aplicar colores fila por fila
    for row in ws.iter_rows(min_row=2, min_col=estado_col, max_col=estado_col):
        for cell in row:
            if cell.value == "Ya vino":
                cell.fill = fill_verde
            elif cell.value == "No vino":
                cell.fill = fill_rojo
            elif cell.value == "Primera vez":
                cell.fill = fill_azul

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Letra de la columna
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar de nuevo en memoria
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    # Respuesta como Excel
    return StreamingResponse(
        final_output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resultados.xlsx"}
    )
